from django.shortcuts import render
from .models import TardinessDetails
from leaves.models import Holidays
from pim.models import Personal_details
from attendance.models import EmployeeAttendance
import calendar
import datetime
# import xlrd
import openpyxl
from django.shortcuts import render,redirect
from .models import TardinessDetails,ShiftDetails,ShiftTimings
from django.http import JsonResponse
# Create your views here.

def policyDetails(request):
    if request.method == 'POST':
        tardinessvalues = TardinessDetails.objects.all()
        if not tardinessvalues:
            tardinessdetails = TardinessDetails(maxallowedlate=request.POST['maxallowedlate'],
                                penalty_late=request.POST['penalty_late'],
                                maxallowedearly=request.POST['maxallowedearly'],
                                penalty_early=request.POST['penalty_early'],
                                holidaycompensatorytype=request.POST['holidaycompensatorytype'],
                                holidaypaymentqty=request.POST['holidaypaymentqty'],
                                weekendcompensatorytype=request.POST['weekendcompensatorytype'],
                                weekendpaymentqty=request.POST['weekendpaymentqty'],
                                holidaydaystoexpire=request.POST['holidaydaystoexpire'],
                                weekenddaystoexpire=request.POST['weekenddaystoexpire'],
                                )
            tardinessdetails.save()
            tardinessvalues = TardinessDetails.objects.all()
            return render(request,'policy.html',{'title':'Attendance Policy Details','tardinessvalues':tardinessvalues})
        else:
            tardinessdetails = TardinessDetails.objects.get(id=1)
            tardinessdetails.maxallowedlate=request.POST['maxallowedlate']
            tardinessdetails.penalty_late=request.POST['penalty_late']
            tardinessdetails.maxallowedearly=request.POST['maxallowedearly']
            tardinessdetails.penalty_early=request.POST['penalty_early']
            tardinessdetails.holidaycompensatorytype=request.POST['holidaycompensatorytype']
            tardinessdetails.holidaypaymentqty=request.POST['holidaypaymentqty']
            tardinessdetails.weekendcompensatorytype=request.POST['weekendcompensatorytype']
            tardinessdetails.weekendpaymentqty=request.POST['weekendpaymentqty']
            tardinessdetails.minimumHoursForOT=request.POST['minimumHoursForOT']
            tardinessdetails.OTcompensatorytype=request.POST['OTcompensatorytype']
            tardinessdetails.OTPayment=request.POST['OTPayment']
            tardinessdetails.holidaydaystoexpire=request.POST['holidaydaystoexpire']
            tardinessdetails.weekenddaystoexpire=request.POST['weekenddaystoexpire']
            tardinessdetails.save()
            tardinessvalues = TardinessDetails.objects.all()
            return render(request,'policy.html',{'title':'Attendance Policy Details','tardinessvalues':tardinessvalues})
    else:
        tardinessvalues = TardinessDetails.objects.all()
        if not tardinessvalues:
            dataflag = False
        else:
            dataflag = True
        return render(request,'policy.html',{'title':'Attendance Policy Details','tardinessvalues':tardinessvalues,'dataflag':dataflag})

def shiftDetails(request):
    if request.method == "GET":
        return render(request,'shiftdetails.html',{'title':'Employee ShiftDetails'})
    else:
        response_data = {}
        shiftname = None
        shiftname = ShiftDetails.objects.filter(shiftname = request.POST['shiftname'])
        if not shiftname:
            response_data["is_success"] = True
            shiftdetail = ShiftDetails(shiftname=request.POST['shiftname'],
                                    shortname=request.POST['shortname'],
                                    shiftdescription=request.POST['shiftdescription'],
                                    flexibleshift=request.POST['flexibleshift'])
            shiftdetail.save()
            shiftid = shiftdetail.id
            mondayshifttiming = ShiftTimings(shift_name_id=shiftid,
                                    workdays= request.POST['mondayworkingday'],
                                    shift_in_time=request.POST['mondayintime'],
                                    shift_out_time=request.POST['mondayouttime'],
                                    from_time=request.POST['mondaybreaktimefrom'],
                                    to_time=request.POST['mondaybreaktimeto'],
                                    weekoff=request.POST['weekoffmonday'])
            tuesdayshifttiming = ShiftTimings(shift_name_id=shiftid,
                                    workdays= request.POST['tuesdayworkingday'],
                                    shift_in_time=request.POST['tuesdayintime'],
                                    shift_out_time=request.POST['tuesdayouttime'],
                                    from_time=request.POST['tuesdaybreaktimefrom'],
                                    to_time=request.POST['tuesdaybreaktimeto'],
                                    weekoff=request.POST['weekofftuesday'])   
            wednesdayshifttiming = ShiftTimings(shift_name_id=shiftid,
                                    workdays= request.POST['wednesdayworkingday'],
                                    shift_in_time=request.POST['wednesdayintime'],
                                    shift_out_time=request.POST['wednesdayouttime'],
                                    from_time=request.POST['wednesdaybreaktimefrom'],
                                    to_time=request.POST['wednesdaybreaktimeto'],
                                    weekoff=request.POST['weekoffwednesday'])  
            thursdayshifttiming = ShiftTimings(shift_name_id=shiftid,
                                    workdays= request.POST['thursdayworkingday'],
                                    shift_in_time=request.POST['thursdayintime'],
                                    shift_out_time=request.POST['thursdayouttime'],
                                    from_time=request.POST['thursdaybreaktimefrom'],
                                    to_time=request.POST['thursdaybreaktimeto'],
                                    weekoff=request.POST['weekoffthursday'])
            fridayshifttiming = ShiftTimings(shift_name_id=shiftid,
                                    workdays= request.POST['fridayworkingday'],
                                    shift_in_time=request.POST['fridayintime'],
                                    shift_out_time=request.POST['fridayouttime'],
                                    from_time=request.POST['fridaybreaktimefrom'],
                                    to_time=request.POST['fridaybreaktimeto'],
                                    weekoff=request.POST['weekofffriday'])
            saturdayshifttiming = ShiftTimings(shift_name_id=shiftid,
                                    workdays= request.POST['saturdayworkingday'],
                                    shift_in_time=request.POST['saturdayintime'],
                                    shift_out_time=request.POST['saturdayouttime'],
                                    from_time=request.POST['saturdaybreaktimefrom'],
                                    to_time=request.POST['saturdaybreaktimeto'],
                                    weekoff=request.POST['weekoffsaturday'])  
            sundayshifttiming = ShiftTimings(shift_name_id=shiftid,
                                    workdays= request.POST['sundayworkingday'],
                                    shift_in_time=request.POST['sundayintime'],
                                    shift_out_time=request.POST['sundayouttime'],
                                    from_time=request.POST['sundaybreaktimefrom'],
                                    to_time=request.POST['sundaybreaktimeto'],
                                    weekoff=request.POST['weekoffsunday'])       
            mondayshifttiming.save()
            tuesdayshifttiming.save()
            wednesdayshifttiming.save()
            thursdayshifttiming.save()
            fridayshifttiming.save()
            saturdayshifttiming.save()
            sundayshifttiming.save()
            return redirect('/attendance/employeeshiftlist')
        else:
            response_data["is_success"] = False
        return JsonResponse(response_data)
    
def validateshiftDetails(request):
    if request.method == "GET":
        return render(request,'shiftdetails.html',{'title':'Employee ShiftDetails'})
    else: 
        response_data = {}
        shortname = None
        shortname = ShiftDetails.objects.filter(shortname = request.POST['shortname'])
        if not shortname:
            response_data["is_success"] = True
            shiftdetail = ShiftDetails(shiftname=request.POST['shiftname'],
                                    shortname=request.POST['shortname'],
                                    shiftdescription=request.POST['shiftdescription'],
                                    flexibleshift=request.POST['flexibleshift'])
            shiftdetail.save()
            return redirect('/attendance/employeeshiftlist')
        else:
            response_data["is_success"] = False
        return JsonResponse(response_data)

def employeeshiftlist(request): 
    shiftdetails = ShiftTimings.objects.all().select_related('shift_name').values('shift_name__shiftname','shift_name__shortname','shift_in_time','shift_out_time').distinct()
    return render(request,'employeeshiftlist.html',{'title':'Employee Shift list','shiftdetails':shiftdetails})

def uploadattendance(request):
    saved_details = {}
    if request.method == "POST":
        new_file = request.FILES["uploadattendance"]
        wb_obj = openpyxl.load_workbook(new_file)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        max_col = sheet_obj.max_column
        # get the attendance date
        cell_obj = sheet_obj.cell(row = 10, column = 6)
        AttendanceDate = cell_obj.value
        AttendanceDate = AttendanceDate.split('-')
        abbr_to_num = {name: num for num, name in enumerate(calendar.month_abbr) if num}
        AttendanceDate[1] = abbr_to_num[AttendanceDate[1]]
        attendanceMarkingDate = str(AttendanceDate[2])+'-'+str(AttendanceDate[1])+'-'+str(AttendanceDate[0])

        # check if Attendancedate is a holiday
        holiday = Holidays.objects.filter(holidayDate__year= AttendanceDate[2],
                                            holidayDate__month= AttendanceDate[1],
                                            holidayDate__day= AttendanceDate[0])
        if not holiday:
            holidaystatus = 2  # not a holiday
        else:
            holidaystatus = 1  # Its a holiday
        
        leavestatus = 2
        compoffstatus = 2

        for i in range(13, max_row):
            Employee_Code = sheet_obj.cell(row = i, column = 4)
            Employee_Name = sheet_obj.cell(row = i, column = 6)
            shift = sheet_obj.cell(row = i, column = 8)
            Intime = sheet_obj.cell(row = i, column = 9)
            outime = sheet_obj.cell(row = i, column = 10)
            workduration = sheet_obj.cell(row = i, column = 13)
            OT = sheet_obj.cell(row = i, column = 14) 
            totalwork = sheet_obj.cell(row = i, column = 16)
            status = sheet_obj.cell(row = i, column = 17)

            Employee_Code = Employee_Code.value
            Employee_Code = str(Employee_Code).zfill(5)
            Employee_Name = Employee_Name.value
            Employee_shift = shift.value
            Employee_Intime = Intime.value
            Employee_outime = outime.value
            Employee_workduration = workduration.value
            Employee_OT = OT.value
            Employee_totalwork = totalwork.value
            Attendance_status = status.value
            personals = Personal_details.objects.filter(employee_id = Employee_Code).values('id','work_shifts')
            
            emp_id = personals[0]['id']
            emp_shift = personals[0]['work_shifts']

            # if day is not holiday ,check if it is a weekend
            if holidaystatus == 2:
                weekendstatus = 2
            else:
                weekendstatus = 2

            # if employee punch-in time is not available
            if Employee_Intime == '00:00':
                # if not a holiday check if is a weekend
                if holidaystatus == 2:
                    #check for week weekend

                    # if not weekend
                    if weekendstatus == 2:
                        #check for leave
                        if leavestatus == 1:
                            attendancestatus = 4   # Leave
                        else:
                            # check for compoff
                            if compoffstatus == 1:
                                attendancestatus = 5   # Comp Off 
                            else:
                                attendancestatus = 2   # Absent
                    else:
                        attendancestatus = 2   # Absent            
                else:
                    attendancestatus = 2   # Absent
            else:
                if holidaystatus == 1:
                    attendancestatus = 7   # Holiday day present
                else:
                    #check for week weekend
                    if weekendstatus == 1:
                        attendancestatus = 6   # Weekend present
                    else:
                        if Attendance_status == 'P':
                            attendancestatus = 1   # Full day present
                        elif Attendance_status == 'A':
                            attendancestatus = 2   # Absent
                        else:
                            #check if leave 
                            if leavestatus == 1:
                                attendancestatus = 3   # Half day present and half day leave 
                            else:
                                attendancestatus = 8   # Half day present and half day Absent              
            
            attendance_details = EmployeeAttendance(employee_name = Employee_Name,
                                                    employee_shift = Employee_shift,
                                                    employee_intime = Employee_Intime,
                                                    employee_outime = Employee_outime,
                                                    employee_workduration = Employee_workduration,
                                                    employee_OT = Employee_OT,
                                                    employee_totalwork = Employee_totalwork,
                                                    holiday_status = holidaystatus,
                                                    weekend_status = weekendstatus,
                                                    attendance_status = attendancestatus,
                                                    leave_status = leavestatus,
                                                    compoff_status = compoffstatus,
                                                    updated_date_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                                    employee_code_id = emp_id,
                                                    attendance_date = attendanceMarkingDate
                                                    )   
            attendance_details.save()
        AttendanceDate = attendanceMarkingDate.split('-')
        saved_details = EmployeeAttendance.objects.filter( attendance_date__year= AttendanceDate[0],attendance_date__month= AttendanceDate[1],attendance_date__day= AttendanceDate[2]).select_related('employee_code').order_by('employee_code_id')
        return render(request,'processattendance.html',{'title':'Upload Attendance', 'saved_details':saved_details})
    else:
        return render(request,'processattendance.html',{'title':'Upload Attendance', 'saved_details':saved_details})



        

